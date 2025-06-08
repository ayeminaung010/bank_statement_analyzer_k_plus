import fitz  # PyMuPDF
import re
from collections import defaultdict
import os
import traceback  # Add this for detailed error tracking
from flask import Flask, request, render_template, redirect, url_for, flash, send_file
import pdfplumber
from datetime import datetime
import pandas as pd
import io
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# NEW: Import the Google AI library
import google.generativeai as genai

# --- Flask Application Setup ---
app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'your-secret-key-here')  # Required for flashing messages

# Create uploads directory if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Configure Gemini AI
GOOGLE_API_KEY = os.getenv('GOOGLE_API_KEY')
if not GOOGLE_API_KEY:
    print("Warning: GOOGLE_API_KEY not found in environment variables")
    model = None
else:
    try:
        genai.configure(api_key=GOOGLE_API_KEY)
        model = genai.GenerativeModel('gemini-1.5-flash')  # Use the fast and efficient Flash model
        print("AI Model configured successfully.")
    except Exception as e:
        print(f"Error configuring AI model: {e}")
        model = None

# --- NEW: AI-powered categorization function ---
def get_category_with_ai(description):
    """Uses the Gemini AI to categorize a transaction description."""
    if not model:
        return "AI Not Configured"

    # A list of your desired categories. This keeps the AI's answers consistent.
    defined_categories = [
        'Groceries', 'Food & Drink', 'Shopping', 'Transportation', 'Bills & Utilities',
        'Digital Wallet Top-up', 'Transfers Out', 'Education', 'Income', 'Uncategorized'
    ]

    # This is the "prompt" we send to the AI.
    prompt = f"""
    You are an expert financial assistant for a user in Thailand.
    Analyze the following bank transaction description and categorize it.
    The description is: "{description}"

    Choose ONLY ONE category from this list: {', '.join(defined_categories)}.
    Return only the category name and nothing else.
    """
    
    try:
        response = model.generate_content(prompt)
        # Clean up the AI's response to get just the category name
        category = response.text.strip()
        # If the AI gives a weird answer, fall back to Uncategorized
        return category if category in defined_categories else 'Uncategorized'
    except Exception as e:
        print(f"AI categorization failed: {e}")
        return "Uncategorized"


# --- PDF Processing Logic ---
def extract_text_from_pdf(pdf_path):
    """Extract text from PDF with focus on table structure preservation."""
    try:
        print(f"\nAttempting to open PDF at: {pdf_path}")
        if not os.path.exists(pdf_path):
            print(f"Error: PDF file not found at {pdf_path}")
            return None

        doc = fitz.open(pdf_path)
        print(f"Successfully opened PDF with {len(doc)} pages")

        all_text = []
        
        for page_num, page in enumerate(doc):
            print(f"\nProcessing page {page_num + 1}")
            
            # Get all text blocks with their coordinates
            blocks = page.get_text("dict")["blocks"]
            
            # Sort blocks by vertical position (top to bottom)
            sorted_blocks = []
            for b in blocks:
                if "lines" in b:
                    for l in b["lines"]:
                        if "spans" in l:
                            for s in l["spans"]:
                                # Get coordinates and text
                                x0, y0, x1, y1 = s["bbox"]
                                text = s["text"].strip()
                                if text:  # Only add non-empty text
                                    sorted_blocks.append((y0, x0, text))
            
            # Sort by y-coordinate (vertical position) first, then x-coordinate
            sorted_blocks.sort()
            
            current_line = []
            current_y = None
            
            # Group text blocks into lines based on their y-coordinates
            for y, x, text in sorted_blocks:
                if current_y is None:
                    current_y = y
                
                # If y position is significantly different, it's a new line
                if abs(y - current_y) > 10:  # Threshold for new line
                    if current_line:
                        all_text.append(" ".join(current_line))
                    current_line = [text]
                    current_y = y
                else:
                    current_line.append(text)
            
            # Add the last line
            if current_line:
                all_text.append(" ".join(current_line))

        doc.close()
        
        # Join all lines with newlines
        final_text = "\n".join(all_text)
        
        print("\n=== Extracted Text Sample ===")
        print("First 500 characters:")
        print(final_text[:500])
        
        return final_text

    except Exception as e:
        print(f"PDF extraction failed: {str(e)}")
        print(f"Full traceback: {traceback.format_exc()}")
        return None

def extract_amount(text):
    """Extract amount from text, handling both withdrawal and deposit formats."""
    if not text or text.isspace():
        return 0.0
    try:
        # Remove any commas and convert to float
        return float(text.replace(',', ''))
    except ValueError:
        return 0.0

def parse_transactions(text):
    """Parse transaction data from the text."""
    transactions = []
    total_withdrawals = 0.0
    total_deposits = 0.0
    withdrawal_count = 0
    deposit_count = 0
    ending_balance = 0.0
    
    # Split text into lines and process each line
    lines = text.split('\n')
    
    # Regular expression for date format DD-MM-YY
    date_pattern = re.compile(r'(\d{2}-\d{2}-\d{2})')
    time_pattern = re.compile(r'(\d{2}:\d{2})')
    
    current_date = None
    for line in lines:
        # Skip empty lines
        if not line.strip():
            continue
            
        # Check for ending balance in the header
        if 'ENDING BALANCE' in line:
            try:
                ending_balance = float(line.split()[-1].replace(',', ''))
            except (ValueError, IndexError):
                pass
            continue

        # Try to parse as a transaction line
        if date_pattern.match(line):
            try:
                # Split the line into parts
                parts = line.split()
                current_date = parts[0]  # Get the date
                
                # Extract time if present
                time = "00:00"
                time_found = False
                description_start = 1  # Start after date by default
                
                # Look for time in the next part
                if len(parts) > 1:
                    time_match = time_pattern.match(parts[1])
                    if time_match:
                        time = parts[1]
                        time_found = True
                        description_start = 2  # Start after time
                
                # Get description
                description_parts = []
                amount = None
                balance = None
                channel = "N/A"
                details = ""
                
                # Process remaining parts
                for i, part in enumerate(parts[description_start:]):
                    try:
                        # Try to convert to float (potential amount/balance)
                        value = float(part.replace(',', ''))
                        if amount is None:
                            amount = value
                        elif balance is None:
                            balance = value
                            # Get channel and details after balance
                            remaining = parts[description_start + i + 1:]
                            if remaining:
                                channel = remaining[0]
                                if len(remaining) > 1:
                                    details = ' '.join(remaining[1:])
                            break
                    except ValueError:
                        if amount is None:  # Still collecting description
                            description_parts.append(part)
                
                description = ' '.join(description_parts)
                
                # Skip "Beginning Balance" entries
                if "Beginning Balance" in description:
                    continue
                
                # Determine if it's a withdrawal or deposit
                is_withdrawal = any(keyword in description.upper() for keyword in ['PAYMENT', 'WITHDRAWAL', 'PAID', 'TO'])
                
                if amount is not None:
                    if is_withdrawal:
                        amount = -abs(amount)
                        total_withdrawals += abs(amount)
                        withdrawal_count += 1
                    else:
                        amount = abs(amount)
                        total_deposits += amount
                        deposit_count += 1

                    # Clean up description
                    description = description.strip()
                    if description.startswith(':'):
                        description = description[1:].strip()
                    
                    # Extract transfer details from channel/details
                    if 'To' in details:
                        transfer_details = details.split('To')[-1].strip()
                        if transfer_details:
                            channel = f"Transfer to {transfer_details}"
                    elif 'From' in details:
                        transfer_details = details.split('From')[-1].strip()
                        if transfer_details:
                            channel = f"Transfer from {transfer_details}"
                    elif 'K PLUS' in channel:
                        if details:
                            channel = f"K PLUS - {details}"

                    transaction = {
                        'date': current_date,
                        'time': time,
                        'description': description,
                        'amount': amount,
                        'balance': balance if balance is not None else 0.0,
                        'channel': channel,
                        'details': details
                    }
                    transactions.append(transaction)

            except Exception as e:
                print(f"Error parsing line: {line}")
                print(f"Error details: {str(e)}")
                continue

    # Sort transactions by date and time
    transactions.sort(key=lambda x: (x['date'], x['time']))
    
    return {
        'transactions': transactions,
        'total_withdrawals': total_withdrawals,
        'total_deposits': total_deposits,
        'withdrawal_count': withdrawal_count,
        'deposit_count': deposit_count,
        'net_amount': total_deposits - total_withdrawals,
        'ending_balance': ending_balance
    }

def analyze_spending(transactions):
    """Analyze spending patterns and categorize transactions."""
    categories = defaultdict(float)
    
    for transaction in transactions:
        if transaction['amount'] < 0:  # Only analyze withdrawals/spending
            amount = abs(transaction['amount'])
            description = transaction['description'].lower()
            
            # Basic categorization logic
            if any(word in description for word in ['food', 'restaurant', 'cafe']):
                categories['Food & Dining'] += amount
            elif any(word in description for word in ['transfer', 'sent']):
                categories['Transfers'] += amount
            elif any(word in description for word in ['transport', 'grab', 'taxi']):
                categories['Transportation'] += amount
            elif any(word in description for word in ['shop', 'store', 'mart']):
                categories['Shopping'] += amount
            elif any(word in description for word in ['bill', 'utility', 'electric', 'water']):
                categories['Bills & Utilities'] += amount
            else:
                categories['Other'] += amount
    
    # Convert to list of tuples and sort by amount
    sorted_categories = sorted(categories.items(), key=lambda x: x[1], reverse=True)
    
    # Prepare data for the chart
    labels = [cat[0] for cat in sorted_categories]
    values = [cat[1] for cat in sorted_categories]
    
    return labels, values

# The rest of your app.py remains the same
# (create_transaction_summary, Flask routes, etc.)

# In your app.py file

@app.route('/')
def index():
    return render_template('index.html')

# Store the last parsed results in memory
last_results = None

def create_excel_file(transactions, summary_data):
    """Create an Excel file with transaction data and summary."""
    # Create a BytesIO object to store the Excel file
    output = io.BytesIO()
    
    # Create Excel writer object
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create transactions dataframe
        df_transactions = pd.DataFrame(transactions)
        
        # Add withdrawal and deposit columns
        df_transactions['withdrawal'] = df_transactions['amount'].apply(
            lambda x: abs(x) if x < 0 else None
        )
        df_transactions['deposit'] = df_transactions['amount'].apply(
            lambda x: x if x > 0 else None
        )
        
        # Reorder and rename columns
        df_transactions = df_transactions[[
            'date', 'time', 'description', 'withdrawal', 'deposit', 'balance', 'channel'
        ]]
        df_transactions.columns = [
            'Date', 'Time', 'Description', 'Withdrawal (THB)', 'Deposit (THB)', 'Balance (THB)', 'Channel/Details'
        ]
        
        # Calculate monthly statistics
        monthly_stats = {
            'Total Withdrawals': df_transactions['Withdrawal (THB)'].sum(),
            'Total Deposits': df_transactions['Deposit (THB)'].sum(),
            'Net Change': df_transactions['Deposit (THB)'].sum() - df_transactions['Withdrawal (THB)'].sum(),
            'Starting Balance': df_transactions.iloc[0]['Balance (THB)'],
            'Ending Balance': df_transactions.iloc[-1]['Balance (THB)'],
            'Number of Withdrawals': df_transactions['Withdrawal (THB)'].count(),
            'Number of Deposits': df_transactions['Deposit (THB)'].count(),
            'Average Withdrawal': df_transactions['Withdrawal (THB)'].mean(),
            'Average Deposit': df_transactions['Deposit (THB)'].mean(),
            'Largest Withdrawal': df_transactions['Withdrawal (THB)'].max(),
            'Largest Deposit': df_transactions['Deposit (THB)'].max()
        }
        
        # Create summary dataframe
        summary_rows = [
            ['Monthly Transaction Summary', ''],
            ['Starting Balance', monthly_stats['Starting Balance']],
            ['Ending Balance', monthly_stats['Ending Balance']],
            ['Net Change', monthly_stats['Net Change']],
            ['', ''],
            ['Withdrawal Summary', ''],
            ['Total Withdrawals', monthly_stats['Total Withdrawals']],
            ['Number of Withdrawals', monthly_stats['Number of Withdrawals']],
            ['Average Withdrawal', monthly_stats['Average Withdrawal']],
            ['Largest Withdrawal', monthly_stats['Largest Withdrawal']],
            ['', ''],
            ['Deposit Summary', ''],
            ['Total Deposits', monthly_stats['Total Deposits']],
            ['Number of Deposits', monthly_stats['Number of Deposits']],
            ['Average Deposit', monthly_stats['Average Deposit']],
            ['Largest Deposit', monthly_stats['Largest Deposit']]
        ]
        
        df_summary = pd.DataFrame(summary_rows, columns=['Category', 'Amount'])
        
        # Write dataframes to Excel
        df_transactions.to_excel(writer, sheet_name='Transactions', index=False)
        
        # Add summary table at the bottom of transactions
        start_row = len(df_transactions) + 3
        summary_start = f'A{start_row}'
        df_summary.to_excel(writer, sheet_name='Transactions', startrow=start_row, index=False)
        
        # Also write summary to separate sheet
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Transactions']
        
        # Import openpyxl styles
        from openpyxl.styles import Font, PatternFill, Color, Border, Side
        from openpyxl.styles.colors import RGB
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 12  # Date
        worksheet.column_dimensions['B'].width = 10  # Time
        worksheet.column_dimensions['C'].width = 40  # Description
        worksheet.column_dimensions['D'].width = 15  # Withdrawal
        worksheet.column_dimensions['E'].width = 15  # Deposit
        worksheet.column_dimensions['F'].width = 15  # Balance
        worksheet.column_dimensions['G'].width = 40  # Channel/Details
        
        # Format number columns (withdrawal, deposit, balance)
        for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=6):
            for cell in row:
                cell.number_format = '#,##0.00'
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0056B3', end_color='0056B3', fill_type='solid')
        withdrawal_font = Font(color='FF0000')  # Red
        deposit_font = Font(color='008000')     # Green
        summary_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Style headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Color code withdrawal and deposit columns
        for row in worksheet.iter_rows(min_row=2):
            withdrawal_cell = row[3]  # Column D
            deposit_cell = row[4]     # Column E
            
            if withdrawal_cell.value:
                withdrawal_cell.font = withdrawal_font
            if deposit_cell.value:
                deposit_cell.font = deposit_font
        
        # Style summary section in transactions sheet
        summary_range = worksheet[f'{summary_start}:B{start_row + len(summary_rows)}']
        for row in summary_range:
            for cell in row:
                cell.border = thin_border
                if cell.column == 1 and cell.value and not cell.value.startswith(' '):  # Headers
                    cell.font = summary_font
                if cell.column == 2:  # Amount column
                    cell.number_format = '#,##0.00'
        
        # Format summary sheet
        summary_sheet = writer.sheets['Summary']
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 15
        
        # Style summary sheet
        for row in summary_sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.column == 1 and cell.value and not cell.value.startswith(' '):  # Headers
                    cell.font = summary_font
                if cell.column == 2:  # Amount column
                    cell.number_format = '#,##0.00'
    
    # Reset pointer to beginning of file
    output.seek(0)
    return output

@app.route('/export_excel')
def export_excel():
    """Export transaction data to Excel."""
    if not last_results:
        return "No data to export. Please analyze a statement first.", 400
        
    try:
        # Create Excel file
        excel_file = create_excel_file(
            last_results['transactions'],
            {
                'total_deposits': last_results['total_deposits'],
                'total_withdrawals': last_results['total_withdrawals'],
                'net_amount': last_results['net_amount'],
                'ending_balance': last_results['ending_balance'],
                'deposit_count': last_results['deposit_count'],
                'withdrawal_count': last_results['withdrawal_count']
            }
        )
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'bank_statement_analysis_{timestamp}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        return f"Error creating Excel file: {str(e)}", 500

@app.route('/analyze', methods=['POST'])
def analyze():
    global last_results
    
    if 'file' not in request.files:
        return render_template('index.html', error='No file uploaded')
    
    file = request.files['file']
    if file.filename == '':
        return render_template('index.html', error='No file selected')
    
    if not file.filename.endswith('.pdf'):
        return render_template('index.html', error='Please upload a PDF file')
    
    try:
        # Ensure upload directory exists
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
        # Generate a unique filename to avoid conflicts
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_filename = f"statement_{timestamp}.pdf"
        filepath = os.path.join(UPLOAD_FOLDER, safe_filename)
        
        # Save the uploaded file
        try:
            file.save(filepath)
            print(f"Saving file to: {filepath}")
        except Exception as save_error:
            print(f"Error saving file: {str(save_error)}")
            return render_template('index.html', error=f"Error saving file: Please try again")
        
        print("Starting PDF text extraction...")
        print(f"\nAttempting to open PDF at: {filepath}")
        
        # Extract text from PDF
        try:
            with pdfplumber.open(filepath) as pdf:
                print(f"Successfully opened PDF with {len(pdf.pages)} pages\n")
                text = ''
                for i, page in enumerate(pdf.pages, 1):
                    print(f"\nProcessing page {i}")
                    text += page.extract_text() + '\n'
        except Exception as pdf_error:
            print(f"Error reading PDF: {str(pdf_error)}")
            return render_template('index.html', error="Error reading PDF file: Please ensure it's a valid bank statement")
        finally:
            # Clean up the uploaded file regardless of success or failure
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
            except Exception as cleanup_error:
                print(f"Warning: Could not remove temporary file {filepath}: {cleanup_error}")
        
        print("\n=== Extracted Text Sample ===")
        print("First 500 characters:")
        print(text[:500])
        
        print("\nStarting transaction parsing...")
        results = parse_transactions(text)
        transactions = results['transactions']
        
        if not transactions:
            return render_template('index.html', error="No transactions found in PDF")
        
        print("Analyzing spending patterns...")
        chart_labels, chart_data = analyze_spending(transactions)
        
        # Add chart data to results
        results['chart_labels'] = chart_labels
        results['chart_data'] = chart_data
        
        # Store results for export
        last_results = results
        
        # Get AI insights if API key is available
        if GOOGLE_API_KEY:
            try:
                model = genai.GenerativeModel('gemini-pro')
                prompt = f"""Analyze these bank transactions and provide insights about spending patterns, unusual transactions, and suggestions for financial management. Focus on key patterns and actionable advice. Keep it concise (max 3-4 bullet points).

Transaction summary:
- Total withdrawals: {results['total_withdrawals']} THB ({results['withdrawal_count']} transactions)
- Total deposits: {results['total_deposits']} THB ({results['deposit_count']} transactions)
- Net amount: {results['net_amount']} THB
- Ending balance: {results['ending_balance']} THB"""

                response = model.generate_content(prompt)
                results['ai_insights'] = response.text
            except Exception as e:
                print(f"Error getting AI insights: {e}")
                results['ai_insights'] = None
        
        return render_template('index.html', results=results, transactions=transactions)
        
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        print(f"Full traceback: {traceback.format_exc()}")
        # Clean up in case of error
        try:
            if 'filepath' in locals() and os.path.exists(filepath):
                os.remove(filepath)
        except Exception as cleanup_error:
            print(f"Warning: Could not remove temporary file during error cleanup: {cleanup_error}")
        return render_template('index.html', error=f"Error processing the PDF file: {str(e)}")

if __name__ == "__main__":
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    app.run(debug=True)